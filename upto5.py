import PySimpleGUI as sg
layout = [[sg.Text('TSEAP')],
[sg.Text('Result Analysis')],
[sg.Text('Comment: ')],
[sg.Text('Please select')],
[sg.Text('      run result files')],
[sg.Text('      below five runs!!')],
[sg.Text('For multiple selecting')],
[sg.Text('Please press Ctrl keys and click!')],
[sg.Cancel()]]

window = sg.Window('Window Title', layout)

event, values = window.Read()
window.Close()

import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilenames()

import sys
mod = sys.modules[__name__]

num= str(len(file_path))
for i in range(int(num)):
    setattr(mod, 'f{}'.format(i), file_path[i])

from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

for i in num:
    if int(i)>0:
        wb1 = load_workbook(f0)
        ws1 = wb1['Axial raw data']

        # define number of data
        x1 = [ws1.cell(row=i, column=1).value for i in range(2, 500)]
        y1 = [ws1.cell(row=i, column=2).value for i in range(2, 500)]

        value_arr_x1 = np.transpose(x1)
        value_arr_y1 = np.transpose(y1)

        df1 = pd.DataFrame(value_arr_x1)
        df_y1 = pd.DataFrame(value_arr_y1)

        # define name of dataframe
        df1[1] = df_y1
        df1.columns = ['x1', 'y1']

        df_filename = pd.DataFrame({'x1': [f0], 'y1': [f0]})

        df_name = pd.DataFrame({})
        # check your file path. you can change 'Axial' as your file name
        df_name['x1'] = df_filename['x1'].str.split('Axial').str[1]
        df_name['y1'] = df_filename['y1'].str.split('Axial').str[1]
        a = df_name.values.T.tolist()
        b = sum(a, [])
        b[0] = b[0] + '_x1'
        b[1] = b[1] + '_y1'


for i in num:
    if int(i)>1:
        wb2 = load_workbook(f1)
        ws2 = wb2['Axial raw data']

        y2 = [ws2.cell(row=i, column=2).value for i in range(2, 500)]
        value_arr_y2 = np.transpose(y2)
        df2 = pd.DataFrame(value_arr_y2)

        # add name of dataframe
        df1['y2'] = df2

        df_filename_f2 = pd.DataFrame({'y2': [f1]})
        df_filename['y2'] =df_filename_f2

        df_name['y2'] = df_filename['y2'].str.split('Axial').str[1]
        a = df_name.values.T.tolist()
        b = sum(a, [])
        b[0] = b[0] + '_x1'
        b[1] = b[1] + '_y1'
        b[2] = b[2] + '_y2'


for i in num:
    if int(i)>2:
        wb3 = load_workbook(f2)
        ws3 = wb3['Axial raw data']

        y3 = [ws3.cell(row=i, column=2).value for i in range(2, 500)]
        value_arr_y3 = np.transpose(y3)
        df3 = pd.DataFrame(value_arr_y3)

        df1['y3'] = df3

        df_filename_f3 = pd.DataFrame({'y3': [f2]})
        df_filename['y3'] = df_filename_f3

        df_name['y3'] = df_filename['y3'].str.split('Axial').str[1]
        a = df_name.values.T.tolist()
        b = sum(a, [])
        b[0] = b[0] + '_x1'
        b[1] = b[1] + '_y1'
        b[2] = b[2] + '_y2'
        b[3] = b[3] + '_y3'

for i in num:
    if int(i)>3:
        wb4 = load_workbook(f3)
        ws4 = wb4['Axial raw data']

        y4 = [ws4.cell(row=i, column=2).value for i in range(2, 500)]
        value_arr_y4 = np.transpose(y4)
        df4 = pd.DataFrame(value_arr_y4)

        df1['y4'] = df4

        df_filename_f4 = pd.DataFrame({'y4': [f3]})
        df_filename['y4'] = df_filename_f4

        df_name['y4'] = df_filename['y4'].str.split('Axial').str[1]
        a = df_name.values.T.tolist()
        b = sum(a, [])
        b[0] = b[0] + '_x1'
        b[1] = b[1] + '_y1'
        b[2] = b[2] + '_y2'
        b[3] = b[3] + '_y3'
        b[4] = b[4] + '_y4'

for i in num:
    if int(i)>4:
        wb5 = load_workbook(f4)
        ws5 = wb5['Axial raw data']

        y5 = [ws5.cell(row=i, column=2).value for i in range(2, 500)]
        value_arr_y5 = np.transpose(y5)
        df5 = pd.DataFrame(value_arr_y5)

        df1['y5'] = df5

        df_filename_f5 = pd.DataFrame({'y5': [f4]})
        df_filename['y5'] = df_filename_f5

        df_name['y5'] = df_filename['y5'].str.split('Axial').str[1]

        a = df_name.values.T.tolist()
        b = sum(a, [])
        b[0] = b[0] + '_x1'
        b[1] = b[1] + '_y1'
        b[2] = b[2] + '_y2'
        b[3] = b[3] + '_y3'
        b[4] = b[4] + '_y4'
        b[5] = b[5] + '_y5'

df1.columns = b

ws_xy = wb1.create_sheet(title= 'xy_sheet')

for row in dataframe_to_rows(df1, index=False, header=True):
  if len(row)>0.1:
    ws_xy.append(row)

# can conunt number of data

from datetime import date
from openpyxl.chart import (
    ScatterChart,
    LineChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis

c1 = LineChart()
# set specifics of chart
c1.title = "Pressure Profile"
c1.style = 10
c1.y_axis.title = 'Pressure [bar]'
c1.x_axis.title = 'Length [mm]'

# define range of data
data = Reference(ws_xy, min_col=2, min_row=1, max_col=6, max_row=320)
c1.add_data(data, titles_from_data=True)

chart = ScatterChart()
chart.title = "Pressure Profile"
chart.style = 10
chart.x_axis.title = 'Length [mm]'
chart.y_axis.title = 'Pressure [bar]'

xvalues = Reference(ws_xy, min_col=1, min_row=2, max_row=320)
for i in range(2, 7):
    values = Reference(ws_xy, min_col=i, min_row=1, max_row=320)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)

# choose location of chart
ws_xy.add_chart(c1, "H15")
ws_xy.add_chart(chart, "H31")


import matplotlib.pyplot as plt
plt.style.use('seaborn-whitegrid')
import numpy as np
fig1 = plt.figure()
ax = fig1.add_subplot()
x_list = []
for i in range(0,498):
    x_list.append(i)
x_number=np.array(x_list)

x = x_number

for i in num:
    if int(i)>0:
        y = value_arr_y1
        ax.plot(x, y)
        plt.plot(x, y)
        plt.xlabel("Length")
        plt.ylabel("Pressure")
        plt.title("Check the ff reference point.")

for i in num:
    if int(i)>1:
        y2 = value_arr_y2
        ax.plot(x, y2)
        plt.plot(x, y,x,y2)
        plt.xlabel("Length")
        plt.ylabel("Pressure")
        plt.title("Check the ff reference point.")

for i in num:
    if int(i)>2:
        y3 = value_arr_y3
        ax.plot(x, y3)
        plt.plot(x, y, x, y2, x, y3)
        plt.xlabel("Length")
        plt.ylabel("Pressure")
        plt.title("Check the ff reference point.")

for i in num:
    if int(i)>3:
        y4 = value_arr_y4
        ax.plot(x, y4)
        plt.plot(x, y, x, y2, x, y3, x, y4)
        plt.xlabel("Length")
        plt.ylabel("Pressure")
        plt.title("Check the ff reference point.")

for i in num:
    if int(i)>4:
        y5 = value_arr_y5
        ax.plot(x, y5)
        plt.plot(x, y, x, y2, x, y3, x, y4, x, y5)
        plt.xlabel("Length")
        plt.ylabel("Pressure")
        plt.title("Check the ff reference point.")

layout = [[sg.Text('Please input a file name for multiselect result!')],
[sg.Text('A file will be created in same directory of "TSEAP kneading"')],
[sg.Text('Do not forget .xlsx')],
[sg.InputText()],
[sg.Submit(), sg.Cancel()]]

window = sg.Window('Window Title', layout)

event, values = window.Read()
window.Close()

file_save = values[0]

wb1.save(file_save)

layout = [[sg.Text('Please find a reference point for filling calculating and press any key')],

[sg.Cancel()]]

window = sg.Window('Window Title', layout)

event, values = window.Read()
window.Close()
plt.show(block=False)
plt.draw()
plt.waitforbuttonpress(0)
plt.close()

layout = [[sg.Text('Please input a reference point')],
[sg.InputText()],
[sg.Submit(), sg.Cancel()]]

window = sg.Window('Window Title', layout)

event, values = window.Read()
window.Close()

insert = int(values[0])
insert_str = str(values[0])

layout_s = [[sg.Text('Please input intermediate reference points up to 4')],
[sg.Text('Even if there are no points. Do not leave blanks. Enter 2 or more numbers.')],
[sg.InputText()],
[sg.InputText()],
[sg.InputText()],
[sg.InputText()],
[sg.Submit(), sg.Cancel()]]

window = sg.Window('Window Title', layout_s)

event, values = window.Read()
window.Close()

insert_1 = str(values[0])
insert_2 = str(values[1])
insert_3 = str(values[2])
insert_4 = str(values[3])

ws1 = wb1['xy_sheet']

o= ws1['B2'].value, ws1['C2'].value, ws1['D2'].value, ws1['E2'].value, ws1['F2']
number=0
for i in range(5):
    if o[i] is not None:
        number +=1

a='A'
b='B'
c='C'
d='D'
e='E'
f='F'
s = '='
m = '-'
x = str(insert+1)
insert_cell_str=str(insert+1)

q15 = '_Q15'
q20 = '_Q20'
q25 = '_Q25'
num = str(number)
for i in num:
    if int(i)>0:
        ws1['I1'] = 'MAX_Value'
        ws1['J1'] = 'T/F'
        ws1['K1'] = 'Cell_Number'
        ws1['H2'] = 'Y1'
        ws1['I2'] ='= MAX(B:B)'
        ws1['J2'] ='=IF(B'+insert_cell_str+'<B2, False, True)'
        ws1['K2']='=MATCH(MAX(B:B),B:B,0)'

        y_Q_15 = [ws1.cell(row=i, column=2).value for i in range(2, 500)]
        value_arr_y1 = np.transpose(y_Q_15)
        df_y1 = pd.DataFrame(value_arr_y1)
        t1 = df_y1.loc[insert - 1:insert - 1, 0:0].values
        df_y1[df_y1 < t1] = 0
        df_y1[df_y1 >= t1] = 1
        df_y1.columns = ['Q15']

        df_y1.loc['Total', :] = df_y1.sum(axis=0)
        y_sum = df_y1.loc['Total', :].values


        y1_start = str(int(insert + 1 - y_sum[0] + 1))

        ws1['M1'] = 'Length_Value'
        ws1['M2'] = 'Y1'
        ws1['N2'] = s + a + x + m + a + y1_start

        ws1['M1'] = 'Length_Value'
        ws1['M2'] = 'Y1'
        ws1['N2'] = s + a + x + m + a + y1_start

        ws1['Q2'] = s + a + insert_1 + m + a + y1_start
        ws1['Q3'] = s + a + insert_2 + m + a + y1_start
        ws1['Q4'] = s + a + insert_3 + m + a + y1_start
        ws1['Q5'] = s + a + insert_4 + m + a + y1_start
        ws1['Q6'] = s + a + x + m + a + y1_start

        ws1['R1'] = 'Pressure difference'
        ws1['R2'] = "=MAX(B:B) -" + b + insert_1
        ws1['R3'] = "=MAX(B:B) -" + b + insert_2
        ws1['R4'] = "=MAX(B:B) -" + b + insert_3
        ws1['R5'] = "=MAX(B:B) -" + b + insert_4
        ws1['R6'] = "=MAX(B:B) -" + b + x

        ws1['H8'] = "Y1="+file_path[0]

for i in num:
    if int(i)>1:
        ws1['H3'] = 'Y2'
        ws1['I3'] ='= MAX(C:C)'
        ws1['J3'] ='=IF(C'+insert_cell_str+'<C2, False, True)'
        ws1['K3']='=MATCH(MAX(C:C),C:C,0)'

        y_Q_20 = [ws1.cell(row=i, column=3).value for i in range(2, 500)]
        value_arr_y2 = np.transpose(y_Q_20)
        df_y2 = pd.DataFrame(value_arr_y2)
        t2 = df_y2.loc[insert - 1:insert - 1, 0:0].values
        df_y2[df_y2 < t2] = 0
        df_y2[df_y2 >= t2] = 1

        df_y2.loc['Total', :] = df_y2.sum(axis=0)
        y_sum_2 = df_y2.loc['Total', :].values

        y2_start = str(int(insert + 1 - y_sum_2[0] + 1))

        ws1['M3'] = 'Y2'
        ws1['N3'] = s + a + x + m + a + y2_start

        ws1['S1'] = 'Y2'
        ws1['S2'] = s + a + insert_1 + m + a + y2_start
        ws1['S3'] = s + a + insert_2 + m + a + y2_start
        ws1['S4'] = s + a + insert_3 + m + a + y2_start
        ws1['S5'] = s + a + insert_4 + m + a + y2_start
        ws1['S6'] = s + a + x + m + a + y2_start

        ws1['T1'] = 'Pressure difference'
        ws1['T2'] = "=MAX(C:C) -" + c + insert_1
        ws1['T3'] = "=MAX(C:C) -" + c + insert_2
        ws1['T4'] = "=MAX(C:C) -" + c + insert_3
        ws1['T5'] = "=MAX(C:C) -" + c + insert_4
        ws1['T6'] = "=MAX(C:C) -" + c + x

        ws1['H9'] = "Y2="+file_path[1]

for i in num:
    if int(i)>2:
        ws1['H4'] = 'Y3'
        ws1['I4'] ='= MAX(D:D)'
        ws1['J4'] ='=IF(D'+insert_cell_str+'<D2, False, True)'
        ws1['K4']='=MATCH(MAX(D:D),D:D,0)'

        y_Q_25 = [ws1.cell(row=i, column=4).value for i in range(2, 500)]
        value_arr_y3 = np.transpose(y_Q_25)
        df_y3 = pd.DataFrame(value_arr_y3)
        t3 = df_y3.loc[insert - 1:insert - 1, 0:0].values
        df_y3[df_y3 < t3] = 0
        df_y3[df_y3 >= t3] = 1

        df_y3.loc['Total', :] = df_y3.sum(axis=0)
        y_sum_3 = df_y3.loc['Total', :].values

        y3_start = str(int(insert + 1 - y_sum_3[0] + 1))

        ws1['M4'] = 'Y3'
        ws1['N4'] = s + a + x + m + a + y3_start

        ws1['U1'] = 'Y3'
        ws1['U2'] = s + a + insert_1 + m + a + y3_start
        ws1['U3'] = s + a + insert_2 + m + a + y3_start
        ws1['U4'] = s + a + insert_3 + m + a + y3_start
        ws1['U5'] = s + a + insert_4 + m + a + y3_start
        ws1['U6'] = s + a + x + m + a + y3_start

        ws1['V1'] = "Pressure difference"
        ws1['V2'] = "=MAX(D:D) -" + d + insert_1
        ws1['V3'] = "=MAX(D:D) -" + d + insert_2
        ws1['V4'] = "=MAX(D:D) -" + d + insert_3
        ws1['V5'] = "=MAX(D:D) -" + d + insert_4
        ws1['V6'] = "=MAX(D:D) -" + d + x

        ws1['H10'] = "Y3="+file_path[2]

for i in num:
    if int(i)>3:
        ws1['H5'] = 'Y4'
        ws1['I5'] ='= MAX(E:E)'
        ws1['J5'] ='=IF(E'+insert_cell_str+'<E2, False, True)'
        ws1['K5']='=MATCH(MAX(E:E),E:E,0)'

        y_4 = [ws1.cell(row=i, column=5).value for i in range(2, 500)]
        value_arr_y4 = np.transpose(y_4)
        df_y4 = pd.DataFrame(value_arr_y4)
        t4 = df_y4.loc[insert - 1:insert - 1, 0:0].values
        df_y4[df_y4 < t4] = 0
        df_y4[df_y4 >= t4] = 1

        df_y4.loc['Total', :] = df_y4.sum(axis=0)
        y_sum_4 = df_y4.loc['Total', :].values

        y4_start = str(int(insert + 1 - y_sum_4[0] + 1))

        ws1['M5'] = 'Y4'
        ws1['N5'] = s + a + x + m + a + y4_start

        ws1['W1'] = 'Y4'
        ws1['W2'] = s + a + insert_1 + m + a + y4_start
        ws1['W3'] = s + a + insert_2 + m + a + y4_start
        ws1['W4'] = s + a + insert_3 + m + a + y4_start
        ws1['W5'] = s + a + insert_4 + m + a + y4_start
        ws1['W6'] = s + a + x + m + a + y4_start

        ws1['X1'] = "Pressure difference"
        ws1['X2'] = "=MAX(E:E) -" + e + insert_1
        ws1['X3'] = "=MAX(E:E) -" + e + insert_2
        ws1['X4'] = "=MAX(E:E) -" + e + insert_3
        ws1['X5'] = "=MAX(E:E) -" + e + insert_4
        ws1['X6'] = "=MAX(E:E) -" + e + x

        ws1['H11'] = "Y4="+file_path[3]

for i in num:
    if int(i)>4:
        ws1['H6'] = 'Y5'
        ws1['I6'] ='= MAX(F:F)'
        ws1['J6'] ='=IF(F'+insert_cell_str+'<F2, False, True)'
        ws1['K6']='=MATCH(MAX(F:F),F:F,0)'

        y_5 = [ws1.cell(row=i, column=6).value for i in range(2, 500)]
        value_arr_y5 = np.transpose(y_5)
        df_y5 = pd.DataFrame(value_arr_y5)
        t5 = df_y5.loc[insert - 1:insert - 1, 0:0].values
        df_y5[df_y5 < t5] = 0
        df_y5[df_y5 >= t5] = 1

        df_y5.loc['Total', :] = df_y5.sum(axis=0)
        y_sum_5 = df_y5.loc['Total', :].values

        y5_start = str(int(insert + 1 - y_sum_5[0] + 1))

        ws1['M6'] = 'Y5'
        ws1['N6'] = s + a + x + m + a + y5_start

        ws1['Y1'] = 'Y5'
        ws1['Y2'] = s + a + insert_1 + m + a + y5_start
        ws1['Y3'] = s + a + insert_2 + m + a + y5_start
        ws1['Y4'] = s + a + insert_3 + m + a + y5_start
        ws1['Y5'] = s + a + insert_4 + m + a + y5_start
        ws1['Y6'] = s + a + x + m + a + y5_start

        ws1['Z1'] = "Pressure difference"
        ws1['Z2'] = "=MAX(F:F) -" + f + insert_1
        ws1['Z3'] = "=MAX(F:F) -" + f + insert_2
        ws1['Z4'] = "=MAX(F:F) -" + f + insert_3
        ws1['Z5'] = "=MAX(F:F) -" + f + insert_4
        ws1['Z6'] = "=MAX(F:F) -" + f + x

        ws1['H12'] = "Y5="+file_path[4]


#--------------------------------------------------------------------------MAX
ws1['P1'] ='Sperated Length_Value'
ws1['P2'] ='5mm'
ws1['P3'] ='10mm'
ws1['P4'] ='15mm'
ws1['P5'] ='20mm'
ws1['P6'] ='25mm'

x1 = [ws1.cell(row=i, column=1).value for i in range(2, 500)]
value_arr_x1 = np.transpose(x1)
df_x1 = pd.DataFrame(value_arr_x1)

for i in num:
    if int(i)>0:
        value_arr_y_a = np.transpose(y_Q_15)
        df_y_a = pd.DataFrame(value_arr_y_a)

for i in num:
    if int(i)>1:
        value_arr_y_b = np.transpose(y_Q_20)
        df_y_b = pd.DataFrame(value_arr_y_b)
        df_y_a[1] = df_y_b

for i in num:
    if int(i)>2:
        value_arr_y_c = np.transpose(y_Q_25)
        df_y_c = pd.DataFrame(value_arr_y_c)
        df_y_a[2] = df_y_c

for i in num:
    if int(i)>3:
        value_arr_y_d = np.transpose(y_4)
        df_y_d = pd.DataFrame(value_arr_y_d)
        df_y_a[3] = df_y_d

for i in num:
    if int(i)>4:
        value_arr_y_e = np.transpose(y_5)
        df_y_e = pd.DataFrame(value_arr_y_e)
        df_y_a[4] = df_y_e

df_y_abc=df_y_a.iloc[insert -1:insert,:]
mah_np_array = df_y_abc.values
ff_list = mah_np_array.tolist()

ff_list_sum = []
for i in range(1,361):
    ff_list_sum +=ff_list
ff_array = np.array(ff_list_sum)
dfre = pd.DataFrame(ff_array)

df_y_sub=df_y_a.sub(dfre)
if number == 1:
    co=['Y1']
if number == 2:
    co=['Y1','Y2']
if number == 3:
    co=['Y1','Y2','Y3']
if number == 4:
    co=['Y1','Y2','Y3','Y4']
if number == 5:
    co=['Y1','Y2','Y3','Y4','Y5']

df_x1[co]= df_y_sub
df_x1 = df_x1.rename(columns = {0: 'x'})

ws_y_sub = wb1.create_sheet(title= 'second_chart')

for row in dataframe_to_rows(df_x1, index=False, header=True):
  if len(row)>0.1:
    ws_y_sub.append(row)


c1 = LineChart()
# set specifics of chart
c1.title = "Pressure Profile"
c1.style = 10
c1.y_axis.title = 'Length [mm]'
c1.x_axis.title = 'Pressure [bar]'

# define range of data
data = Reference(ws_y_sub, min_col=2, min_row=1, max_col=6, max_row=362)
c1.add_data(data, titles_from_data=True)

chart = ScatterChart()
chart.title = "Pressure Profile"
chart.style = 10
chart.x_axis.title = 'Pressure [bar]'
chart.y_axis.title = 'Length [mm]'

xvalues = Reference(ws_y_sub, min_col=1, min_row=2, max_row=361)
for i in range(2, 7):
    values = Reference(ws_y_sub, min_col=i, min_row=1, max_row=361)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)

# choose location of chart
ws1.add_chart(c1, "P15")
ws1.add_chart(chart, "P31")

fig2 = plt.figure()
ax2 = fig2.add_subplot()
for i in num:
    if int(i)>0:
        x_number = [ws1.cell(row=i, column=1).value for i in range(2, 500)]
        x_n = np.transpose(x_number)
        y1_number = [ws1.cell(row=i, column=2).value for i in range(2, 500)]
        y_f = np.transpose(y1_number)
        ax2.plot(x_n, y_f)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'

for i in num:
    if int(i)>1:
        y2_number = [ws1.cell(row=i, column=3).value for i in range(2, 500)]
        y_s = np.transpose(y2_number)
        ax2.plot(x_n, y_s)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>2:
        y3_number = [ws1.cell(row=i, column=4).value for i in range(2, 500)]
        y_t = np.transpose(y3_number)
        ax2.plot(x_n, y_t)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>3:
        y4_number = [ws1.cell(row=i, column=5).value for i in range(2, 500)]
        y_fo = np.transpose(y4_number)
        ax2.plot(x_n, y_fo)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>4:
        y5_number = [ws1.cell(row=i, column=6).value for i in range(2, 500)]
        y_fi = np.transpose(y5_number)
        ax2.plot(x_n, y_fi)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
plt.show(block=False)
plt.draw()
plt.waitforbuttonpress(0)
plt.close()

fig3 = plt.figure()
ax3 = fig3.add_subplot()

for i in num:
    if int(i)>0:
        y_mf = df_x1['Y1'].values
        ax3.plot(x_n, y_mf)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'

for i in num:
    if int(i)>1:
        y_ms = df_x1['Y2'].values
        ax3.plot(x_n, y_ms)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>2:
        y_mt = df_x1['Y3'].values
        ax3.plot(x_n, y_mt)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>3:
        y_mfo = df_x1['Y4'].values
        ax3.plot(x_n, y_mfo)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
for i in num:
    if int(i)>4:
        y_mfi = df_x1['Y5'].values
        ax3.plot(x_n, y_mfi)
        plt.xlabel = 'Length'
        plt.ylabel = 'Pressure'
plt.draw()
plt.waitforbuttonpress(0)
plt.close()

for i in num:
    if int(i)>0:
        x_s = pd.DataFrame(x_n)
        x_s[co]=df_y_a
        x_s = x_s.rename(columns = {0: 'x'})
        insert_fi=int(insert_1) -2
        insert_se=int(insert_2) -2
        insert_th=int(insert_3) -2
        insert_fo=int(insert_4) -2
        insert_to=insert-1
        y1_start_int=int(y1_start)-2

        qwe_1=x_s.iloc[insert_fi:insert_fi+1,0:1].values -x_s.iloc[y1_start_int:y1_start_int+1,0:1].values
        qwe_2=x_s.iloc[insert_se:insert_se+1,0:1].values -x_s.iloc[y1_start_int:y1_start_int+1,0:1].values
        qwe_3=x_s.iloc[insert_th:insert_th+1,0:1].values -x_s.iloc[y1_start_int:y1_start_int+1,0:1].values
        qwe_4=x_s.iloc[insert_fo:insert_fo+1,0:1].values -x_s.iloc[y1_start_int:y1_start_int+1,0:1].values
        qwe_5=x_s.iloc[insert_to:insert_to+1,0:1].values -x_s.iloc[y1_start_int:y1_start_int+1,0:1].values

        df_qwe1=pd.DataFrame(qwe_1)
        df_qwe1[1]=pd.DataFrame(qwe_2)
        df_qwe1[2]=pd.DataFrame(qwe_3)
        df_qwe1[3]=pd.DataFrame(qwe_4)
        df_qwe1[4]=pd.DataFrame(qwe_5)

        df_qwe=df_qwe1.T

        df_asd1 = pd.DataFrame(y_f)
        asd_1=df_asd1.max().values-df_asd1.iloc[insert_fi:insert_fi+1,0:1].values
        asd_2=df_asd1.max().values-df_asd1.iloc[insert_se:insert_se+1,0:1].values
        asd_3=df_asd1.max().values-df_asd1.iloc[insert_th:insert_th+1,0:1].values
        asd_4=df_asd1.max().values-df_asd1.iloc[insert_fo:insert_fo+1,0:1].values
        asd_5=df_asd1.max().values-df_asd1.iloc[insert_to:insert_to+1,0:1].values

        df_asd2=pd.DataFrame(asd_1)
        df_asd2[1]=pd.DataFrame(asd_2)
        df_asd2[2]=pd.DataFrame(asd_3)
        df_asd2[3]=pd.DataFrame(asd_4)
        df_asd2[4]=pd.DataFrame(asd_5)
        df_asd=df_asd2.T

        df_qwe[1]=df_asd
        df_qwe=df_qwe.rename(columns={0:'Y1'})
        df_qwe = df_qwe.rename(columns={1: 'Pressure difference_Y1'})

for i in num:
    if int(i)>1:
        y2_start_int=int(y2_start)-2

        rty_1=x_s.iloc[insert_fi:insert_fi+1,0:1].values -x_s.iloc[y2_start_int:y2_start_int+1,0:1].values
        rty_2=x_s.iloc[insert_se:insert_se+1,0:1].values -x_s.iloc[y2_start_int:y2_start_int+1,0:1].values
        rty_3=x_s.iloc[insert_th:insert_th+1,0:1].values -x_s.iloc[y2_start_int:y2_start_int+1,0:1].values
        rty_4=x_s.iloc[insert_fo:insert_fo+1,0:1].values -x_s.iloc[y2_start_int:y2_start_int+1,0:1].values
        rty_5=x_s.iloc[insert_to:insert_to+1,0:1].values -x_s.iloc[y2_start_int:y2_start_int+1,0:1].values

        df_rty1=pd.DataFrame(rty_1)
        df_rty1[1]=pd.DataFrame(rty_2)
        df_rty1[2]=pd.DataFrame(rty_3)
        df_rty1[3]=pd.DataFrame(rty_4)
        df_rty1[4]=pd.DataFrame(rty_5)

        df_rty = df_rty1.T

        df_fgh1 = pd.DataFrame(y_s)
        fgh_1 = df_fgh1.max().values - df_fgh1.iloc[insert_fi:insert_fi + 1, 0:1].values
        fgh_2 = df_fgh1.max().values - df_fgh1.iloc[insert_se:insert_se + 1, 0:1].values
        fgh_3 = df_fgh1.max().values - df_fgh1.iloc[insert_th:insert_th + 1, 0:1].values
        fgh_4 = df_fgh1.max().values - df_fgh1.iloc[insert_fo:insert_fo + 1, 0:1].values
        fgh_5 = df_fgh1.max().values - df_fgh1.iloc[insert_to:insert_to + 1, 0:1].values

        df_fgh2 = pd.DataFrame(fgh_1)
        df_fgh2[1] = pd.DataFrame(fgh_2)
        df_fgh2[2] = pd.DataFrame(fgh_3)
        df_fgh2[3] = pd.DataFrame(fgh_4)
        df_fgh2[4] = pd.DataFrame(fgh_5)
        df_fgh = df_fgh2.T

        df_qwe['Y2']=df_rty
        df_qwe['Pressure difference_Y2']=df_fgh

for i in num:
    if int(i)>2:
        y3_start_int = int(y3_start) - 2

        uio_1 = x_s.iloc[insert_fi:insert_fi + 1, 0:1].values - x_s.iloc[y3_start_int:y3_start_int + 1, 0:1].values
        uio_2 = x_s.iloc[insert_se:insert_se + 1, 0:1].values - x_s.iloc[y3_start_int:y3_start_int + 1, 0:1].values
        uio_3 = x_s.iloc[insert_th:insert_th + 1, 0:1].values - x_s.iloc[y3_start_int:y3_start_int + 1, 0:1].values
        uio_4 = x_s.iloc[insert_fo:insert_fo + 1, 0:1].values - x_s.iloc[y3_start_int:y3_start_int + 1, 0:1].values
        uio_5 = x_s.iloc[insert_to:insert_to + 1, 0:1].values - x_s.iloc[y3_start_int:y3_start_int + 1, 0:1].values

        df_uio1 = pd.DataFrame(uio_1)
        df_uio1[1] = pd.DataFrame(uio_2)
        df_uio1[2] = pd.DataFrame(uio_3)
        df_uio1[3] = pd.DataFrame(uio_4)
        df_uio1[4] = pd.DataFrame(uio_5)

        df_uio = df_uio1.T

        df_jkl1 = pd.DataFrame(y_t)
        jkl_1 = df_jkl1.max().values - df_jkl1.iloc[insert_fi:insert_fi + 1, 0:1].values
        jkl_2 = df_jkl1.max().values - df_jkl1.iloc[insert_se:insert_se + 1, 0:1].values
        jkl_3 = df_jkl1.max().values - df_jkl1.iloc[insert_th:insert_th + 1, 0:1].values
        jkl_4 = df_jkl1.max().values - df_jkl1.iloc[insert_fo:insert_fo + 1, 0:1].values
        jkl_5 = df_jkl1.max().values - df_jkl1.iloc[insert_to:insert_to + 1, 0:1].values

        df_jkl2 = pd.DataFrame(jkl_1)
        df_jkl2[1] = pd.DataFrame(jkl_2)
        df_jkl2[2] = pd.DataFrame(jkl_3)
        df_jkl2[3] = pd.DataFrame(jkl_4)
        df_jkl2[4] = pd.DataFrame(jkl_5)
        df_jkl = df_jkl2.T

        df_qwe['Y3']=df_uio
        df_qwe['Pressure difference_Y3']=df_jkl

for i in num:
    if int(i)>3:
        y4_start_int = int(y4_start) - 2

        thr_1 = x_s.iloc[insert_fi:insert_fi + 1, 0:1].values - x_s.iloc[y4_start_int:y4_start_int + 1, 0:1].values
        thr_2 = x_s.iloc[insert_se:insert_se + 1, 0:1].values - x_s.iloc[y4_start_int:y4_start_int + 1, 0:1].values
        thr_3 = x_s.iloc[insert_th:insert_th + 1, 0:1].values - x_s.iloc[y4_start_int:y4_start_int + 1, 0:1].values
        thr_4 = x_s.iloc[insert_fo:insert_fo + 1, 0:1].values - x_s.iloc[y4_start_int:y4_start_int + 1, 0:1].values
        thr_5 = x_s.iloc[insert_to:insert_to + 1, 0:1].values - x_s.iloc[y4_start_int:y4_start_int + 1, 0:1].values

        df_thr1 = pd.DataFrame(thr_1)
        df_thr1[1] = pd.DataFrame(thr_2)
        df_thr1[2] = pd.DataFrame(thr_3)
        df_thr1[3] = pd.DataFrame(thr_4)
        df_thr1[4] = pd.DataFrame(thr_5)

        df_thr = df_thr1.T

        df_athr1 = pd.DataFrame(y_fo)
        athr_1 = df_athr1.max().values - df_athr1.iloc[insert_fi:insert_fi + 1, 0:1].values
        athr_2 = df_athr1.max().values - df_athr1.iloc[insert_se:insert_se + 1, 0:1].values
        athr_3 = df_athr1.max().values - df_athr1.iloc[insert_th:insert_th + 1, 0:1].values
        athr_4 = df_athr1.max().values - df_athr1.iloc[insert_fo:insert_fo + 1, 0:1].values
        athr_5 = df_athr1.max().values - df_athr1.iloc[insert_to:insert_to + 1, 0:1].values

        df_athr2 = pd.DataFrame(athr_1)
        df_athr2[1] = pd.DataFrame(athr_2)
        df_athr2[2] = pd.DataFrame(athr_3)
        df_athr2[3] = pd.DataFrame(athr_4)
        df_athr2[4] = pd.DataFrame(athr_5)
        df_athr = df_athr2.T

        df_qwe['Y4']=df_athr
        df_qwe['Pressure difference_Y4']=df_athr

for i in num:
    if int(i) > 4:
        y5_start_int = int(y5_start) - 2

        com_1 = x_s.iloc[insert_fi:insert_fi + 1, 0:1].values - x_s.iloc[y5_start_int:y5_start_int + 1, 0:1].values
        com_2 = x_s.iloc[insert_se:insert_se + 1, 0:1].values - x_s.iloc[y5_start_int:y5_start_int + 1, 0:1].values
        com_3 = x_s.iloc[insert_th:insert_th + 1, 0:1].values - x_s.iloc[y5_start_int:y5_start_int + 1, 0:1].values
        com_4 = x_s.iloc[insert_fo:insert_fo + 1, 0:1].values - x_s.iloc[y5_start_int:y5_start_int + 1, 0:1].values
        com_5 = x_s.iloc[insert_to:insert_to + 1, 0:1].values - x_s.iloc[y5_start_int:y5_start_int + 1, 0:1].values

        df_com1 = pd.DataFrame(com_1)
        df_com1[1] = pd.DataFrame(com_2)
        df_com1[2] = pd.DataFrame(com_3)
        df_com1[3] = pd.DataFrame(com_4)
        df_com1[4] = pd.DataFrame(com_5)

        df_com = df_com1.T

        df_bot1 = pd.DataFrame(y_fi)
        bot_1 = df_bot1.max().values - df_bot1.iloc[insert_fi:insert_fi + 1, 0:1].values
        bot_2 = df_bot1.max().values - df_bot1.iloc[insert_se:insert_se + 1, 0:1].values
        bot_3 = df_bot1.max().values - df_bot1.iloc[insert_th:insert_th + 1, 0:1].values
        bot_4 = df_bot1.max().values - df_bot1.iloc[insert_fo:insert_fo + 1, 0:1].values
        bot_5 = df_bot1.max().values - df_bot1.iloc[insert_to:insert_to + 1, 0:1].values

        df_bot2 = pd.DataFrame(bot_1)
        df_bot2[1] = pd.DataFrame(bot_2)
        df_bot2[2] = pd.DataFrame(bot_3)
        df_bot2[3] = pd.DataFrame(bot_4)
        df_bot2[4] = pd.DataFrame(bot_5)
        df_bot = df_bot2.T

        df_qwe['Y5'] = df_bot
        df_qwe['Pressure difference_Y5'] = df_bot

pd.set_option('display.max_rows', 10)
pd.set_option('display.max_columns', 10)
pd.set_option('display.width', 100)



layout = [[sg.Text('Please enter a file name. Do not forget .xlsx')],
[sg.Text('Length values')],
[sg.Text(str(df_qwe))],
[sg.InputText()],
[sg.Submit(), sg.Cancel()]]

window = sg.Window('Window Title', layout)

event, values = window.Read()
window.Close()

file_save = values[0]
wb1.save(file_save)
